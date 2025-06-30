from openpyxl import Workbook
from openpyxl.drawing.image import Image as ExcelImage
from openpyxl.styles import Alignment, Font
from decimal import Decimal
from PIL import Image
from io import BytesIO
import os

def generar_excel_retiro_raee_bytes(datos, titulo, filtro):
    wb = Workbook()
    ws = wb.active
    ws.title = "Reporte RAEE"

    ws.column_dimensions["A"].width = 10
    ws.column_dimensions["B"].width = 60
    ws.column_dimensions["C"].width = 15

    fila = 1

    # Imagen de cabecera
    img = ExcelImage("Aplicaciones/Producto/cabecera.png")
    img.width = 590
    img.height = 130
    ws.add_image(img, f"A{fila}")
    fila += 6

    # Título
    ws.merge_cells(start_row=fila, start_column=1, end_row=fila, end_column=4)
    titulo_cell = ws.cell(row=fila, column=1, value=titulo)
    titulo_cell.font = Font(bold=True)
    titulo_cell.alignment = Alignment(horizontal="center", vertical="center")
    fila += 1

    # Cabecera
    ws.cell(row=fila, column=1, value="ITEM").font = Font(bold=True)
    ws.cell(row=fila, column=1).alignment = Alignment(horizontal="center", vertical="center")

    ws.cell(row=fila, column=2, value=filtro).font = Font(bold=True)
    ws.cell(row=fila, column=2).alignment = Alignment(horizontal="center", vertical="center")

    ws.cell(row=fila, column=3, value="KGS").font = Font(bold=True)
    ws.cell(row=fila, column=3).alignment = Alignment(horizontal="center", vertical="center")

    fila += 1

    total = 0
    for i, d in enumerate(datos, start=1):
        ws.cell(row=fila, column=1, value=i)
        ws.cell(row=fila, column=2, value=d["Datos"])
        peso = float(d["total_peso"])
        ws.cell(row=fila, column=3, value=peso)
        total += peso
        fila += 1

    # Total general
    ws.cell(row=fila, column=2, value="TOTAL GENERAL").font = Font(bold=True)
    ws.cell(row=fila, column=3, value=total).font = Font(bold=True)
    fila += 2

    # Imagen de pie
    img = ExcelImage("Aplicaciones/Producto/Fotter.png")
    img.width = 590
    img.height = 70
    ws.add_image(img, f"A{fila}")

    # Guardar en memoria
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output
